import os
import time
import requests
import textwrap
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import langextract as lx

# CONFIG
OPENAI_API_KEY = "sk-proj-Y_e8HGSSO1ScDT3BlbkFJTtBBAjbGJ7B53QrlV8ladFQjFq84kZ1dwXpDMsf2TkAuois7MXGLi0Dl7pvZgEW9XBZOSP458A"
MODEL = "gpt-4o"
OUTPUT_FILENAME = "article_keywords.xlsx"

# Save next to the script file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BASE_DIR, OUTPUT_FILENAME)

# 1) fetch article text
def fetch_article_text(url: str) -> str:
    res = requests.get(url, timeout=15)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "html.parser")

    # remove noisy tags
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    paragraphs = [p.get_text(separator=" ", strip=True) for p in soup.find_all("p")]
    return "\n".join(paragraphs)


# 2) extract keywords via langextract
def extract_keywords(text: str):
    prompt = textwrap.dedent("""
        Extract important keywords and phrases from the text.
        Each extraction should be an exact text span — no paraphrasing.
        Add a meaningful attribute 'category' describing the keyword type (e.g. 'technology', 'person', 'concept', etc).
    """)

    examples = [
        lx.data.ExampleData(
            text=(
                "Elasticsearch is a distributed, open-source search and analytics engine "
                "that helps you store, search, and analyze large volumes of data quickly."
            ),
            extractions=[
                lx.data.Extraction(
                    extraction_class="keyword",
                    extraction_text="Elasticsearch",
                    attributes={"category": "technology"},
                ),
                lx.data.Extraction(
                    extraction_class="keyword",
                    extraction_text="open-source search",
                    attributes={"category": "concept"},
                ),
                lx.data.Extraction(
                    extraction_class="keyword",
                    extraction_text="analytics engine",
                    attributes={"category": "technology"},
                ),
            ],
        )
    ]

    result = lx.extract(
        text_or_documents=text,
        prompt_description=prompt,
        examples=examples,
        model_id=MODEL,
        api_key=OPENAI_API_KEY,
        fence_output=True,
        use_schema_constraints=False
    )

    rows = []
    for e in result.extractions:
        if e.extraction_class == "keyword":
            rows.append({
                "keyword": e.extraction_text,
                "category": e.attributes.get("category", "")
            })
    return rows


# 3) URL groups (correct lists)
articles = {
    "Marathi": [
        "https://mahamoney.com/unclaimed-money-check-claim-process-bank-insurance",
        "https://mahamoney.com/investment-slow-progress-7-year-rule-compound-interest",
        "https://mahamoney.com/avadhut-sathe-sebi-crackdown-546-crore-impounded-ban",
        "https://mahamoney.com/demat-account-nomination-digital-process-step-by-step",
        "https://mahamoney.com/fixed-deposit-investment-tax-rules-tds-itr-important"
    ],
    "English": [
        "https://english.mahamoney.com/bima-vistaar-affordable-insurance-for-your-familys-peace-of-mind",
        "https://english.mahamoney.com/life-insurance-payout-for-spouse-protect-your-family-from-creditors",
        "https://english.mahamoney.com/govt-plans-e-commerce-export-boost-for-msmes",
        "https://english.mahamoney.com/indias-plastic-industry-a-double-edged-sword-unwrapping-growth-pollution-solutions",
        "https://english.mahamoney.com/how-agentic-ai-payments-are-transforming-digital-transactions-1"
    ],
    "Hindi": [
        "https://hindi.mahamoney.com/ai-workslop-myths-reality-economic-losses-and-the-road-to-recovery",
        "https://hindi.mahamoney.com/from-begging-bowl-to-bread-basket-role-of-m-s-swaminathan-c-subramaniam-and-b-sivaraman",
        "https://hindi.mahamoney.com/severity-of-financial-fraud-in-india-estimates-prevention-and-technological-solutions",
        "https://hindi.mahamoney.com/indias-art-market-from-record-auctions-to-new-heights",
        "https://hindi.mahamoney.com/mutual-fund-total-expense-ratio-what-is-it-and-how-does-it-affect-your-earnings"
    ]
}


# 4) merge URL cells (vertical) in Excel
def merge_url_cells(excel_path: str):
    wb = load_workbook(excel_path)
    # iterate over each sheet and perform merging per sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_url = None
        start_row = None

        # header at row 1, so start from 2
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=1)  # column A

            # treat None and "" same for grouping
            val = cell.value if cell.value is not None else ""

            if val != current_url:
                # close previous block
                if current_url is not None and start_row is not None and (r - start_row) > 1:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=r-1, end_column=1)
                    merged_cell = ws.cell(row=start_row, column=1)
                    merged_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

                current_url = val
                start_row = r

        # final block close (if last rows are same)
        last = ws.max_row
        if current_url is not None and start_row is not None and (last - start_row + 1) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=last, end_column=1)
            merged_cell = ws.cell(row=start_row, column=1)
            merged_cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    wb.save(excel_path)


# 5) MAIN: iterate languages -> urls -> extract -> write
def main():
    # accumulate and write per-language sheets
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        for lang, links in articles.items():
            print(f"\nProcessing language: {lang}  (total links: {len(links)})")
            rows = []
            for url in links:
                try:
                    print("  ->", url)
                    article_text = fetch_article_text(url)
                except Exception as e:
                    print(f"     ! fetch failed for {url}: {e}")
                    continue

                try:
                    kw_rows = extract_keywords(article_text)
                except Exception as e:
                    print(f"     ! extraction failed for {url}: {e}")
                    kw_rows = []

                if not kw_rows:
                    # still add a row so the URL is visible (with empty keyword)
                    rows.append({"url": url, "keyword": "", "category": ""})
                else:
                    for r in kw_rows:
                        rows.append({"url": url, "keyword": r.get("keyword", ""), "category": r.get("category", "")})

                # polite pause to avoid hitting API rate limits
                time.sleep(0.5)

            df = pd.DataFrame(rows, columns=["url", "keyword", "category"])
            # ensure sheet name length is valid for Excel
            sheet_name = lang[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # after writing file, merge repeated URL cells per sheet
    merge_url_cells(OUTPUT_PATH)
    print("\n✅ Excel created at:", OUTPUT_PATH)


if __name__ == "__main__":
    main()